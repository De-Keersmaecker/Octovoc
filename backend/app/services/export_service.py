import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from app.models.user import User
from app.models.classroom import Classroom
from app.models.module import Module
from app.models.progress import StudentProgress, QuestionProgress
from datetime import datetime
import os


class ExportService:
    @staticmethod
    def export_classroom_to_excel(classroom_id):
        """Export classroom progress to Excel file"""
        classroom = Classroom.query.get(classroom_id)
        if not classroom:
            raise ValueError('Classroom not found')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Classroom Progress'

        # Headers
        headers = ['Student', 'Module', 'Voortgang (%)', 'Score (%)', 'Juiste antwoorden', 'Totaal vragen', 'Voltooiingsdatum']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Get students
        students = User.query.filter_by(classroom_id=classroom_id).all()

        for student in students:
            progress_records = StudentProgress.query.filter_by(user_id=student.id).all()

            for progress in progress_records:
                module = Module.query.get(progress.module_id)

                # Get total words in module
                from app.models.module import Word
                total_words = Word.query.filter_by(module_id=module.id).count()

                # Calculate score and track answered words
                total_questions = 0
                correct_questions = 0
                answered_word_ids = set()

                for bp in progress.battery_progress:
                    questions = QuestionProgress.query.filter_by(battery_progress_id=bp.id).all()
                    total_questions += len(questions)
                    correct_questions += len([q for q in questions if q.is_correct])

                    # Track unique words answered
                    for q in questions:
                        if q.word_id:
                            answered_word_ids.add(q.word_id)

                answered_words = len(answered_word_ids)

                # Calculate completion percentage (based on unique words answered)
                completion_percentage = (answered_words / total_words * 100) if total_words > 0 else 0

                # Calculate score (based on correct answers)
                score = (correct_questions / total_questions * 100) if total_questions > 0 else 0

                ws.append([
                    student.email,
                    module.name,
                    round(completion_percentage, 1),
                    round(score, 1),
                    correct_questions,
                    total_questions,
                    progress.completion_date.strftime('%Y-%m-%d') if progress.completion_date else 'Bezig'
                ])

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save file
        filename = f'classroom_{classroom_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # Create uploads directory if it doesn't exist
        upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
        os.makedirs(upload_dir, exist_ok=True)

        filepath = os.path.join(upload_dir, filename)
        wb.save(filepath)

        return filepath

    @staticmethod
    def export_student_to_pdf(student_id):
        """Export detailed student report to PDF"""
        student = User.query.get(student_id)
        if not student:
            raise ValueError('Student not found')

        # Create PDF
        filename = f'student_{student_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        # Create uploads directory if it doesn't exist
        upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
        os.makedirs(upload_dir, exist_ok=True)

        filepath = os.path.join(upload_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.black,
            spaceAfter=30,
        )
        heading_style = styles['Heading2']
        normal_style = styles['Normal']

        # Title
        title = Paragraph(f'Student Progress Report: {student.email}', title_style)
        story.append(title)
        story.append(Spacer(1, 0.5 * cm))

        # Student info
        info = [
            ['Email:', student.email],
            ['Role:', student.role],
            ['Registered:', student.created_at.strftime('%Y-%m-%d') if student.created_at else 'N/A'],
            ['Last Activity:', student.last_activity.strftime('%Y-%m-%d') if student.last_activity else 'N/A']
        ]

        info_table = Table(info, colWidths=[4 * cm, 12 * cm])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 1 * cm))

        # Module progress
        story.append(Paragraph('Module Progress', heading_style))
        story.append(Spacer(1, 0.5 * cm))

        progress_records = StudentProgress.query.filter_by(user_id=student_id).all()

        if not progress_records:
            story.append(Paragraph('No progress records found.', normal_style))
        else:
            for progress in progress_records:
                module = Module.query.get(progress.module_id)

                # Module name
                story.append(Paragraph(f'<b>{module.name}</b>', normal_style))

                # Calculate statistics
                total_questions = 0
                correct_questions = 0
                incorrect_words = []

                for bp in progress.battery_progress:
                    questions = QuestionProgress.query.filter_by(battery_progress_id=bp.id).all()
                    total_questions += len(questions)

                    for q in questions:
                        if q.is_correct:
                            correct_questions += 1
                        else:
                            incorrect_words.append(q.word.word if q.word else 'Unknown')

                score = (correct_questions / total_questions * 100) if total_questions > 0 else 0

                # Progress table
                progress_data = [
                    ['Score:', f'{round(score, 2)}%'],
                    ['Questions:', f'{correct_questions}/{total_questions}'],
                    ['Status:', 'Completed' if progress.is_completed else 'In Progress'],
                    ['Completion Date:', progress.completion_date.strftime('%Y-%m-%d') if progress.completion_date else 'N/A']
                ]

                progress_table = Table(progress_data, colWidths=[4 * cm, 10 * cm])
                progress_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
                    ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 9),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(progress_table)

                # Incorrect words
                if incorrect_words:
                    story.append(Spacer(1, 0.3 * cm))
                    story.append(Paragraph(f'<b>Incorrect words:</b> {", ".join(set(incorrect_words[:10]))}', normal_style))

                story.append(Spacer(1, 0.7 * cm))

        # Build PDF
        doc.build(story)

        return filepath
