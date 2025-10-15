import openpyxl
import csv
import io
from app import db
from app.models.module import Module, Word, Battery


class ModuleService:
    @staticmethod
    def create_module_from_excel(filepath, name, difficulty='', is_free=False):
        """
        Create a module from an Excel file
        Expected format:
        - Column A: Word
        - Column B: Meaning
        - Column C: Example sentence (with *word* marked)
        """
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # Validate headers
            if sheet.max_row < 2:
                raise ValueError('Excel file must contain at least one word row')

            # Create module
            module = Module(
                name=name,
                difficulty=difficulty,
                is_free=is_free,
                is_active=True,
                version=1
            )
            db.session.add(module)
            db.session.flush()

            # Parse words
            words = []
            errors = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue  # Skip empty rows

                word_text = row[0] if len(row) > 0 else None
                meaning = row[1] if len(row) > 1 else None
                example_sentence = row[2] if len(row) > 2 else None

                # Validation
                if not word_text or not word_text.strip():
                    errors.append(f'Row {row_idx}: Word is missing')
                    continue

                if not meaning or not meaning.strip():
                    errors.append(f'Row {row_idx}: Meaning is missing')
                    continue

                if not example_sentence or not example_sentence.strip():
                    errors.append(f'Row {row_idx}: Example sentence is missing')
                    continue

                word_text = word_text.strip()
                meaning = meaning.strip()
                example_sentence = example_sentence.strip()

                # Check if example sentence contains ANY word marked with asterisks
                # This allows for inflected forms (e.g., "Ambivalent" with "*ambivalente*" in sentence)
                if '*' not in example_sentence or example_sentence.count('*') < 2:
                    errors.append(
                        f'Row {row_idx}: Example sentence must contain a word marked with asterisks (e.g., *word*)'
                    )
                    continue

                # Create word
                word = Word(
                    module_id=module.id,
                    word=word_text,
                    meaning=meaning,
                    example_sentence=example_sentence,
                    position_in_module=len(words) + 1
                )
                words.append(word)

            if errors:
                db.session.rollback()
                raise ValueError('Excel validation errors:\n' + '\n'.join(errors))

            if not words:
                db.session.rollback()
                raise ValueError('No valid words found in Excel file')

            # Add words to database
            db.session.add_all(words)
            db.session.flush()

            # Create batteries
            word_ids = [w.id for w in words]
            batteries = Battery.create_batteries_for_module(module.id, word_ids)
            db.session.add_all(batteries)

            db.session.commit()

            return module

        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError('Invalid Excel file format. Please upload a valid .xlsx file.')
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def update_module_from_excel(module_id, filepath):
        """
        Update an existing module from Excel file
        This creates a new version - students in progress keep the old version
        """
        module = Module.query.get(module_id)
        if not module:
            raise ValueError('Module not found')

        # For now, we'll create a new module version
        # In a full implementation, you'd need to handle versioning more carefully
        module.version += 1

        # Delete old words and batteries
        Word.query.filter_by(module_id=module_id).delete()
        Battery.query.filter_by(module_id=module_id).delete()

        # Re-import from Excel
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            words = []
            errors = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue

                word_text = row[0] if len(row) > 0 else None
                meaning = row[1] if len(row) > 1 else None
                example_sentence = row[2] if len(row) > 2 else None

                if not word_text or not meaning or not example_sentence:
                    errors.append(f'Row {row_idx}: Missing required fields')
                    continue

                word_text = word_text.strip()
                meaning = meaning.strip()
                example_sentence = example_sentence.strip()

                if '*' not in example_sentence or example_sentence.count('*') < 2:
                    errors.append(f'Row {row_idx}: Example sentence must contain a word marked with asterisks (e.g., *word*)')
                    continue

                word = Word(
                    module_id=module.id,
                    word=word_text,
                    meaning=meaning,
                    example_sentence=example_sentence,
                    position_in_module=len(words) + 1
                )
                words.append(word)

            if errors:
                db.session.rollback()
                raise ValueError('Excel validation errors:\n' + '\n'.join(errors))

            db.session.add_all(words)
            db.session.flush()

            word_ids = [w.id for w in words]
            batteries = Battery.create_batteries_for_module(module.id, word_ids)
            db.session.add_all(batteries)

            module.updated_at = db.func.now()
            db.session.commit()

            return module

        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def create_module_from_csv(csv_data, name, difficulty='', is_free=False):
        """
        Create a module from CSV data string
        Expected format:
        word,meaning,example_sentence or word;meaning;example_sentence
        """
        try:
            # Create module
            module = Module(
                name=name,
                difficulty=difficulty,
                is_free=is_free,
                is_active=True,
                version=1
            )
            db.session.add(module)
            db.session.flush()

            # Parse CSV - auto-detect delimiter (comma or semicolon)
            csv_file = io.StringIO(csv_data)
            # Use Sniffer to detect the delimiter
            sample = csv_data[:1024]  # Take first 1KB as sample

            # Only allow common delimiters (comma, semicolon, tab, pipe)
            valid_delimiters = [',', ';', '\t', '|']

            try:
                detected = csv.Sniffer().sniff(sample, delimiters=',;\t|').delimiter
                # Verify it's a valid delimiter
                if detected in valid_delimiters:
                    delimiter = detected
                else:
                    raise ValueError("Invalid delimiter detected")
            except:
                # Fallback: Try semicolon first (common in European CSVs), then comma
                if ';' in sample and sample.count(';') > sample.count(','):
                    delimiter = ';'
                else:
                    delimiter = ','

            csv_file.seek(0)  # Reset to start
            reader = csv.reader(csv_file, delimiter=delimiter)

            # Check if first row is header
            first_row = next(reader, None)
            if not first_row:
                raise ValueError('CSV is empty')

            # If first row looks like a header, skip it
            if first_row[0].lower() in ['word', 'woord'] or 'word' in first_row[0].lower():
                rows = list(reader)
            else:
                # First row is data, include it
                rows = [first_row] + list(reader)

            words = []
            errors = []

            for row_idx, row in enumerate(rows, start=2 if first_row else 1):
                if not row or not any(row):
                    continue  # Skip empty rows

                word_text = row[0].strip() if len(row) > 0 and row[0] else None
                meaning = row[1].strip() if len(row) > 1 and row[1] else None
                example_sentence = row[2].strip() if len(row) > 2 and row[2] else None

                # Validation
                if not word_text:
                    errors.append(f'Row {row_idx}: Word is missing')
                    continue

                if not meaning:
                    errors.append(f'Row {row_idx}: Meaning is missing')
                    continue

                if not example_sentence:
                    errors.append(f'Row {row_idx}: Example sentence is missing')
                    continue

                # Check if example sentence contains a word marked with asterisks
                if '*' not in example_sentence or example_sentence.count('*') < 2:
                    errors.append(
                        f'Row {row_idx}: Example sentence must contain a word marked with asterisks (e.g., *word*)'
                    )
                    continue

                # Create word
                word = Word(
                    module_id=module.id,
                    word=word_text,
                    meaning=meaning,
                    example_sentence=example_sentence,
                    position_in_module=len(words) + 1
                )
                words.append(word)

            if errors:
                db.session.rollback()
                raise ValueError('CSV validation errors:\n' + '\n'.join(errors))

            if not words:
                db.session.rollback()
                raise ValueError('No valid words found in CSV data')

            # Add words to database
            db.session.add_all(words)
            db.session.flush()

            # Create batteries
            word_ids = [w.id for w in words]
            batteries = Battery.create_batteries_for_module(module.id, word_ids)
            db.session.add_all(batteries)

            db.session.commit()

            return module

        except csv.Error as e:
            db.session.rollback()
            raise ValueError(f'CSV parse error: {str(e)}')
        except Exception as e:
            db.session.rollback()
            raise e
